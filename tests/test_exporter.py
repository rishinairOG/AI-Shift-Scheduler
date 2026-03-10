from datetime import date
import openpyxl
from io import BytesIO
from src.models.staff import Staff
from src.models.schedule import ScheduleConfig
from src.exporter.excel_exporter import export_schedule


def test_export_produces_valid_xlsx():
    staff_list = [
        Staff("bar_1", "Alice", "BAR", "Bartender"),
        Staff("bar_2", "Bob", "BAR", "Bartender"),
    ]
    sections = {"BAR": staff_list}
    config = ScheduleConfig(
        restaurant_name="Test",
        week_start_day="SUN",
        off_request_deadline="FRI",
        publish_day="SAT",
        start_date=date(2026, 3, 8),
        end_date=date(2026, 3, 14),
        active_shift_slots=["C", "D"],
        section_min_per_day={},
        section_max_off_per_day={},
    )
    assignments = {
        ("bar_1", date(2026, 3, 8)): "C",
        ("bar_1", date(2026, 3, 9)): "OFF",
        ("bar_2", date(2026, 3, 8)): "D",
        ("bar_2", date(2026, 3, 9)): "C",
    }
    output = export_schedule(staff_list, sections, config, assignments)
    assert isinstance(output, BytesIO)
    wb = openpyxl.load_workbook(output)
    assert len(wb.sheetnames) > 0


def test_export_contains_staff_names():
    staff_list = [Staff("bar_1", "Alice", "BAR", "Bartender")]
    sections = {"BAR": staff_list}
    config = ScheduleConfig(
        restaurant_name="Test",
        week_start_day="SUN",
        off_request_deadline="FRI",
        publish_day="SAT",
        start_date=date(2026, 3, 8),
        end_date=date(2026, 3, 14),
        active_shift_slots=["C"],
        section_min_per_day={},
        section_max_off_per_day={},
    )
    assignments = {("bar_1", date(2026, 3, 8)): "C"}
    output = export_schedule(staff_list, sections, config, assignments)
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    names_in_sheet = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
    assert "Alice" in names_in_sheet
