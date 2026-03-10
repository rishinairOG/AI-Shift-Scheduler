import openpyxl
from pathlib import Path
from src.models.staff import Staff


SECTION_KEYWORDS = {
    "MANAGERS", "ASST. MANAGERS", "SUPERVISORS", "SERVERS",
    "GRE", "BAR", "Housekeeping", "HOUSEKEEPING",
}


def _normalize_section(name: str) -> str:
    return name.strip().upper()


def parse_roster(file_path: str) -> tuple[list[Staff], dict[str, list[Staff]]]:
    """Parse an Excel roster file and return staff list and section mapping."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_name = "dec" if "dec" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    staff_list: list[Staff] = []
    sections: dict[str, list[Staff]] = {}
    current_section: str | None = None
    section_counter: dict[str, int] = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        col_a = row[0] if len(row) > 0 else None
        col_b = row[1] if len(row) > 1 else None

        if col_b is None:
            continue

        col_b_str = str(col_b).strip()

        # Detect section header
        is_section_header = (
            col_b_str.upper() in {k.upper() for k in SECTION_KEYWORDS}
            and (col_a is None or str(col_a).strip().upper() == "S.NO")
        )
        if is_section_header:
            current_section = col_b_str
            if current_section not in sections:
                sections[current_section] = []
                section_counter[current_section] = 0
            continue

        # Detect staff row
        if current_section and col_a is not None:
            try:
                int(float(str(col_a)))
            except (ValueError, TypeError):
                continue

            name = col_b_str
            if not name:
                continue

            section_counter[current_section] += 1
            staff_id = f"{current_section.lower().replace(' ', '_').replace('.', '')}_{section_counter[current_section]}"
            designation = _infer_designation(current_section)

            staff = Staff(
                id=staff_id,
                name=name,
                section=current_section,
                designation=designation,
            )
            staff_list.append(staff)
            sections[current_section].append(staff)

    return staff_list, sections


def _infer_designation(section: str) -> str:
    mapping = {
        "MANAGERS": "Manager",
        "ASST. MANAGERS": "Asst. Manager",
        "SUPERVISORS": "Supervisor",
        "SERVERS": "Server",
        "GRE": "GRE",
        "BAR": "Bartender",
        "Housekeeping": "Housekeeping",
        "HOUSEKEEPING": "Housekeeping",
    }
    return mapping.get(section, section)
