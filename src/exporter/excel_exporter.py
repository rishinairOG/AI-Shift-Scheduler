from io import BytesIO
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.models.staff import Staff
from src.models.schedule import ScheduleConfig


FILL_OFF = PatternFill("solid", fgColor="C0C0C0")       # grey
FILL_PH = PatternFill("solid", fgColor="FFFF99")        # yellow
FILL_OVERRIDE = PatternFill("solid", fgColor="FFD580")  # light orange
FILL_SHIFT = PatternFill("solid", fgColor="FFFFFF")     # white
FILL_SECTION = PatternFill("solid", fgColor="D9E1F2")   # light blue for section headers

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

DAY_ABBREVS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def _get_fill(code: str) -> PatternFill:
    if code == "OFF":
        return FILL_OFF
    if code in ("PH", "HP"):
        return FILL_PH
    if code in ("S", "HD", "HO", "TR", "V"):
        return FILL_OVERRIDE
    return FILL_SHIFT


def export_schedule(
    staff_list: list[Staff],
    sections: dict[str, list[Staff]],
    config: ScheduleConfig,
    assignments: dict[tuple[str, date], str],
) -> BytesIO:
    """Generate Excel schedule matching the existing roster format."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = config.start_date.strftime("%b").lower()

    # Build date list
    dates = []
    current = config.start_date
    while current <= config.end_date:
        dates.append(current)
        current += timedelta(days=1)

    n_dates = len(dates)
    header_cols = 2  # col A = S.NO, col B = NAMES

    # Row 1: year, month, date numbers
    ws.cell(row=1, column=1, value=config.start_date.year)
    ws.cell(row=1, column=2, value=config.start_date.strftime("%B").upper())
    for j, d in enumerate(dates):
        cell = ws.cell(row=1, column=header_cols + 1 + j, value=d.day)
        cell.alignment = Alignment(horizontal="center")

    # Row 2: "S.NO", "NAMES", day abbreviations
    ws.cell(row=2, column=1, value="S.NO")
    ws.cell(row=2, column=2, value="NAMES")
    for j, d in enumerate(dates):
        cell = ws.cell(row=2, column=header_cols + 1 + j, value=DAY_ABBREVS[d.weekday()])
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True)

    # Staff rows grouped by section
    current_row = 3
    section_order = list(sections.keys())

    for sec_name in section_order:
        sec_staff = sections[sec_name]
        if not sec_staff:
            continue

        # Section header row
        ws.cell(row=current_row, column=1, value="S.NO")
        sec_cell = ws.cell(row=current_row, column=2, value=sec_name)
        sec_cell.font = Font(bold=True)
        sec_cell.fill = FILL_SECTION
        ws.cell(row=current_row, column=1).fill = FILL_SECTION
        current_row += 1

        # Staff rows
        for s_idx, staff in enumerate(sec_staff):
            ws.cell(row=current_row, column=1, value=s_idx + 1)
            ws.cell(row=current_row, column=2, value=staff.name)
            for j, d in enumerate(dates):
                code = assignments.get((staff.id, d), "")
                cell = ws.cell(row=current_row, column=header_cols + 1 + j, value=code)
                cell.fill = _get_fill(code)
                cell.alignment = Alignment(horizontal="center")
                cell.border = THIN_BORDER
            current_row += 1

        # Blank separator row
        current_row += 1

    # Summary rows
    current_row += 1
    ws.cell(row=current_row, column=2, value="Total Working").font = Font(bold=True)
    for j, d in enumerate(dates):
        total = sum(
            1 for s in staff_list
            if assignments.get((s.id, d), "") not in ("OFF", "PH", "HP", "S", "HD", "HO", "TR", "V", "")
        )
        ws.cell(row=current_row, column=header_cols + 1 + j, value=total)
    current_row += 1

    ws.cell(row=current_row, column=2, value="Total OFF").font = Font(bold=True)
    for j, d in enumerate(dates):
        total_off = sum(1 for s in staff_list if assignments.get((s.id, d), "") == "OFF")
        cell = ws.cell(row=current_row, column=header_cols + 1 + j, value=total_off)
        cell.fill = FILL_OFF

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    for j in range(n_dates):
        ws.column_dimensions[get_column_letter(header_cols + 1 + j)].width = 7

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
